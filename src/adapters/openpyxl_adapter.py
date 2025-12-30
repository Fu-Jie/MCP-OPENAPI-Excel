"""
Openpyxl adapter for Excel reading and writing.

This module provides the OpenpyxlAdapter class that wraps openpyxl
for reading and writing Excel files. Openpyxl is a pure Python library
that provides comprehensive Excel support including formulas, styles,
and charts.

Use cases where openpyxl is preferred:
    - When you need to preserve or modify existing Excel formatting
    - When working with formulas that need to be evaluated
    - When python-calamine or XlsxWriter cannot handle specific features
    - When you need to modify existing workbooks (not just create new ones)

Supported formats:
    - .xlsx (Excel 2007+)
    - .xlsm (Excel Macro-Enabled)

Example:
    adapter = OpenpyxlAdapter()

    # Read operations
    sheet_names = adapter.get_sheet_names("/path/to/file.xlsx")
    data = adapter.read_sheet("/path/to/file.xlsx", "Sheet1")

    # Write operations
    adapter.write_sheet(
        "/path/to/output.xlsx",
        rows=[["Name", "Age"], ["Alice", 30], ["Bob", 25]],
        sheet_name="Users"
    )
"""

import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.exceptions.excel_exceptions import (
    CellRangeError,
    InvalidFileFormatError,
    ReadError,
    SheetNotFoundError,
    WriteError,
)
from src.exceptions.excel_exceptions import FileNotFoundError as ExcelFileNotFoundError
from src.exceptions.excel_exceptions import (
    PermissionError as ExcelPermissionError,
)
from src.models.excel_models import CellRange, SheetData, SheetInfo, WorkbookInfo


class OpenpyxlAdapter:
    """
    Adapter for openpyxl Excel reading and writing operations.

    This adapter provides both read and write capabilities using openpyxl.
    It is particularly useful when you need to:
    - Preserve or modify existing Excel formatting
    - Work with formulas
    - Modify existing workbooks
    - Handle edge cases that other adapters cannot

    Attributes:
        SUPPORTED_EXTENSIONS: Tuple of supported file extensions.
        DEFAULT_COLUMN_WIDTH: Default column width in characters.
        MAX_COLUMN_WIDTH: Maximum column width in characters.

    Example:
        adapter = OpenpyxlAdapter()

        # Read operations
        info = adapter.get_workbook_info("/path/to/file.xlsx")
        data = adapter.read_sheet("/path/to/file.xlsx", sheet_name="Data")
        data = adapter.read_range("/path/to/file.xlsx", "A1:C10", sheet_name="Data")

        # Write operations
        result = adapter.write_sheet(
            "/path/to/output.xlsx",
            rows=[["Alice", 30], ["Bob", 25]],
            headers=["Name", "Age"]
        )
    """

    SUPPORTED_EXTENSIONS = (".xlsx", ".xlsm")
    DEFAULT_COLUMN_WIDTH = 10
    MAX_COLUMN_WIDTH = 50

    def __init__(self) -> None:
        """Initialize the OpenpyxlAdapter."""
        pass

    def _validate_file_path(self, file_path: str) -> Path:
        """
        Validate that the file exists and has a supported extension.

        Args:
            file_path: Path to the Excel file.

        Returns:
            Path object for the validated file.

        Raises:
            ExcelFileNotFoundError: If the file does not exist.
            InvalidFileFormatError: If the file extension is not supported.
        """
        path = Path(file_path)

        if not path.exists():
            raise ExcelFileNotFoundError(file_path)

        if path.suffix.lower() not in self.SUPPORTED_EXTENSIONS:
            raise InvalidFileFormatError(
                file_path=file_path,
                expected_formats=list(self.SUPPORTED_EXTENSIONS),
                reason=f"Unsupported file extension: {path.suffix}",
            )

        return path

    def _validate_output_path(
        self,
        file_path: str,
        overwrite: bool = False,
    ) -> Path:
        """
        Validate and prepare the output file path.

        Args:
            file_path: Path where the file will be written.
            overwrite: Whether to overwrite if the file exists.

        Returns:
            Path object for the output file.

        Raises:
            WriteError: If the file exists and overwrite is False.
            ExcelPermissionError: If the directory is not writable.
        """
        path = Path(file_path)

        if path.exists() and not overwrite:
            raise WriteError(
                file_path=file_path,
                operation="create",
                reason="File already exists and overwrite is False",
            )

        parent = path.parent
        if not parent.exists():
            try:
                parent.mkdir(parents=True, exist_ok=True)
            except PermissionError as e:
                raise ExcelPermissionError(
                    file_path=str(parent),
                    operation="create directory",
                ) from e
            except Exception as e:
                raise WriteError(
                    file_path=file_path,
                    operation="create directory",
                    reason=str(e),
                ) from e

        if parent.exists() and not os.access(str(parent), os.W_OK):
            raise ExcelPermissionError(
                file_path=file_path,
                operation="write",
            )

        if not file_path.lower().endswith((".xlsx", ".xlsm")):
            path = Path(file_path + ".xlsx")

        return path

    def _open_workbook(self, file_path: str, data_only: bool = True) -> Workbook:
        """
        Open an Excel workbook using openpyxl.

        Args:
            file_path: Path to the Excel file.
            data_only: If True, read cell values instead of formulas.

        Returns:
            Workbook instance.

        Raises:
            InvalidFileFormatError: If the file cannot be parsed.
            ReadError: If an unexpected error occurs during opening.
        """
        path = self._validate_file_path(file_path)

        try:
            return load_workbook(str(path), data_only=data_only)
        except Exception as e:
            error_msg = str(e).lower()
            if "invalid" in error_msg or "corrupt" in error_msg or "format" in error_msg:
                raise InvalidFileFormatError(
                    file_path=file_path,
                    reason=str(e),
                ) from e
            raise ReadError(
                file_path=file_path,
                operation="open",
                reason=str(e),
            ) from e

    def _normalize_cell_value(self, cell: Cell) -> Any:
        """
        Normalize a cell value from openpyxl to Python types.

        Handles conversion of Excel-specific types to standard Python types.

        Args:
            cell: Cell object from openpyxl.

        Returns:
            Normalized Python value.
        """
        value = cell.value

        if value is None:
            return None

        if isinstance(value, float):
            if value == int(value):
                return int(value)
            return value

        if isinstance(value, (str, int, bool, datetime)):
            return value

        return str(value)

    def _parse_a1_notation(self, a1_range: str) -> CellRange:
        """
        Parse A1 notation range string into CellRange.

        Supports formats like:
        - "A1" (single cell)
        - "A1:C10" (range)

        Args:
            a1_range: A1 notation string.

        Returns:
            CellRange object.

        Raises:
            CellRangeError: If the range notation is invalid.
        """
        a1_range = a1_range.strip().upper()

        single_cell_pattern = r"^([A-Z]+)(\d+)$"
        range_pattern = r"^([A-Z]+)(\d+):([A-Z]+)(\d+)$"

        single_match = re.match(single_cell_pattern, a1_range)
        if single_match:
            col = self._column_letter_to_index(single_match.group(1))
            row = int(single_match.group(2)) - 1
            return CellRange(
                start_row=row,
                end_row=row,
                start_col=col,
                end_col=col,
                a1_notation=a1_range,
            )

        range_match = re.match(range_pattern, a1_range)
        if range_match:
            start_col = self._column_letter_to_index(range_match.group(1))
            start_row = int(range_match.group(2)) - 1
            end_col = self._column_letter_to_index(range_match.group(3))
            end_row = int(range_match.group(4)) - 1

            if start_row > end_row or start_col > end_col:
                raise CellRangeError(
                    cell_range=a1_range,
                    reason="Start position must be before end position",
                )

            return CellRange(
                start_row=start_row,
                end_row=end_row,
                start_col=start_col,
                end_col=end_col,
                a1_notation=a1_range,
            )

        raise CellRangeError(
            cell_range=a1_range,
            reason="Invalid A1 notation format. Expected format: 'A1' or 'A1:C10'",
        )

    def _column_letter_to_index(self, column_letter: str) -> int:
        """
        Convert Excel column letter(s) to 0-based index.

        Args:
            column_letter: Column letter(s) like "A", "B", "AA", "AB".

        Returns:
            0-based column index.
        """
        result = 0
        for char in column_letter.upper():
            result = result * 26 + (ord(char) - ord("A") + 1)
        return result - 1

    def _parse_start_cell(self, start_cell: str) -> tuple[int, int]:
        """
        Parse A1 notation to row and column indices.

        Args:
            start_cell: Cell reference in A1 notation (e.g., "A1", "B5").

        Returns:
            Tuple of (row_index, col_index), both 0-based.
        """
        match = re.match(r"^([A-Za-z]+)(\d+)$", start_cell.strip())
        if not match:
            return (0, 0)

        col_letters = match.group(1).upper()
        row_num = int(match.group(2))

        col_index = 0
        for char in col_letters:
            col_index = col_index * 26 + (ord(char) - ord("A") + 1)
        col_index -= 1

        row_index = row_num - 1

        return (row_index, col_index)

    def _calculate_column_widths(
        self,
        rows: list[list[Any]],
        headers: list[str] | None = None,
    ) -> list[int]:
        """
        Calculate optimal column widths based on content.

        Args:
            rows: Data rows.
            headers: Optional header row.

        Returns:
            List of column widths.
        """
        all_rows = rows.copy()
        if headers:
            all_rows.insert(0, headers)

        if not all_rows:
            return []

        max_cols = max(len(row) for row in all_rows)
        widths = [self.DEFAULT_COLUMN_WIDTH] * max_cols

        for row in all_rows:
            for col_idx, cell in enumerate(row):
                if cell is not None:
                    cell_str = str(cell)
                    cell_width = min(len(cell_str) + 2, self.MAX_COLUMN_WIDTH)
                    widths[col_idx] = max(widths[col_idx], cell_width)

        return widths

    # ==================== READ OPERATIONS ====================

    def get_sheet_names(self, file_path: str) -> list[str]:
        """
        Get the list of sheet names in the workbook.

        Args:
            file_path: Path to the Excel file.

        Returns:
            List of sheet names.

        Raises:
            ExcelFileNotFoundError: If the file does not exist.
            InvalidFileFormatError: If the file format is not supported.
        """
        workbook = self._open_workbook(file_path)
        sheet_names = workbook.sheetnames
        workbook.close()
        return sheet_names

    def get_workbook_info(self, file_path: str) -> WorkbookInfo:
        """
        Get metadata about an Excel workbook.

        Args:
            file_path: Path to the Excel file.

        Returns:
            WorkbookInfo containing workbook metadata.

        Raises:
            ExcelFileNotFoundError: If the file does not exist.
            InvalidFileFormatError: If the file format is not supported.
        """
        path = self._validate_file_path(file_path)
        workbook = self._open_workbook(file_path)

        sheet_names = workbook.sheetnames
        sheets: list[SheetInfo] = []

        for index, name in enumerate(sheet_names):
            try:
                worksheet = workbook[name]
                row_count = worksheet.max_row
                column_count = worksheet.max_column
            except Exception:
                row_count = None
                column_count = None

            sheets.append(
                SheetInfo(
                    name=name,
                    index=index,
                    visible=True,
                    row_count=row_count,
                    column_count=column_count,
                )
            )

        workbook.close()

        stat = path.stat()

        return WorkbookInfo(
            file_path=str(path.absolute()),
            file_size_bytes=stat.st_size,
            sheet_count=len(sheet_names),
            sheets=sheets,
            modified_at=datetime.fromtimestamp(stat.st_mtime),
        )

    def read_sheet(
        self,
        file_path: str,
        sheet_name: str | None = None,
        sheet_index: int | None = None,
        skip_empty_rows: bool = False,
    ) -> SheetData:
        """
        Read data from a specific sheet.

        Args:
            file_path: Path to the Excel file.
            sheet_name: Name of the sheet to read. If None and sheet_index is None,
                       reads the first sheet.
            sheet_index: Index of the sheet to read (0-based). Used if sheet_name is None.
            skip_empty_rows: Whether to skip empty rows.

        Returns:
            SheetData containing the sheet contents.

        Raises:
            ExcelFileNotFoundError: If the file does not exist.
            InvalidFileFormatError: If the file format is not supported.
            SheetNotFoundError: If the specified sheet does not exist.
        """
        workbook = self._open_workbook(file_path)
        available_sheets = workbook.sheetnames

        if sheet_name is not None:
            if sheet_name not in available_sheets:
                workbook.close()
                raise SheetNotFoundError(
                    sheet_name=sheet_name,
                    available_sheets=available_sheets,
                )
            target_sheet_name = sheet_name
        elif sheet_index is not None:
            if sheet_index < 0 or sheet_index >= len(available_sheets):
                workbook.close()
                raise SheetNotFoundError(
                    sheet_name=f"index {sheet_index}",
                    available_sheets=available_sheets,
                )
            target_sheet_name = available_sheets[sheet_index]
        else:
            if not available_sheets:
                workbook.close()
                raise SheetNotFoundError(
                    sheet_name="(first sheet)",
                    available_sheets=[],
                )
            target_sheet_name = available_sheets[0]

        try:
            worksheet = workbook[target_sheet_name]
            rows: list[list[Any]] = []

            for row in worksheet.iter_rows():
                normalized_row = [self._normalize_cell_value(cell) for cell in row]

                if skip_empty_rows:
                    if all(cell is None or cell == "" for cell in normalized_row):
                        continue

                rows.append(normalized_row)

            workbook.close()

        except Exception as e:
            workbook.close()
            raise ReadError(
                file_path=file_path,
                operation="read sheet",
                reason=str(e),
            ) from e

        column_count = max(len(row) for row in rows) if rows else 0

        return SheetData(
            sheet_name=target_sheet_name,
            rows=rows,
            row_count=len(rows),
            column_count=column_count,
        )

    def read_range(
        self,
        file_path: str,
        cell_range: str,
        sheet_name: str | None = None,
        sheet_index: int | None = None,
    ) -> SheetData:
        """
        Read a specific cell range from a sheet.

        Args:
            file_path: Path to the Excel file.
            cell_range: Cell range in A1 notation (e.g., "A1:C10").
            sheet_name: Name of the sheet to read. If None, reads the first sheet.
            sheet_index: Index of the sheet to read (0-based). Used if sheet_name is None.

        Returns:
            SheetData containing the range contents.

        Raises:
            ExcelFileNotFoundError: If the file does not exist.
            InvalidFileFormatError: If the file format is not supported.
            SheetNotFoundError: If the specified sheet does not exist.
            CellRangeError: If the cell range is invalid.
        """
        parsed_range = self._parse_a1_notation(cell_range)

        full_sheet = self.read_sheet(
            file_path=file_path,
            sheet_name=sheet_name,
            sheet_index=sheet_index,
        )

        start_row = parsed_range.start_row
        end_row = min(parsed_range.end_row, full_sheet.row_count - 1)
        start_col = parsed_range.start_col
        end_col = parsed_range.end_col

        if start_row > full_sheet.row_count - 1:
            return SheetData(
                sheet_name=full_sheet.sheet_name,
                rows=[],
                row_count=0,
                column_count=0,
                cell_range=parsed_range,
            )

        extracted_rows: list[list[Any]] = []
        for row_idx in range(start_row, end_row + 1):
            if row_idx < len(full_sheet.rows):
                row = full_sheet.rows[row_idx]
                row_slice = row[start_col : end_col + 1]

                while len(row_slice) < (end_col - start_col + 1):
                    row_slice.append(None)

                extracted_rows.append(row_slice)
            else:
                extracted_rows.append([None] * (end_col - start_col + 1))

        return SheetData(
            sheet_name=full_sheet.sheet_name,
            rows=extracted_rows,
            row_count=len(extracted_rows),
            column_count=end_col - start_col + 1,
            cell_range=parsed_range,
        )

    def get_cell_value(
        self,
        file_path: str,
        cell: str,
        sheet_name: str | None = None,
        sheet_index: int | None = None,
    ) -> Any:
        """
        Get the value of a single cell.

        Args:
            file_path: Path to the Excel file.
            cell: Cell reference in A1 notation (e.g., "A1").
            sheet_name: Name of the sheet. If None, uses the first sheet.
            sheet_index: Index of the sheet (0-based). Used if sheet_name is None.

        Returns:
            The cell value (can be None, str, int, float, bool, or datetime).

        Raises:
            ExcelFileNotFoundError: If the file does not exist.
            InvalidFileFormatError: If the file format is not supported.
            SheetNotFoundError: If the specified sheet does not exist.
            CellRangeError: If the cell reference is invalid.
        """
        range_data = self.read_range(
            file_path=file_path,
            cell_range=cell,
            sheet_name=sheet_name,
            sheet_index=sheet_index,
        )

        if range_data.rows and range_data.rows[0]:
            return range_data.rows[0][0]

        return None

    # ==================== WRITE OPERATIONS ====================

    def write_sheet(
        self,
        file_path: str,
        rows: list[list[Any]],
        sheet_name: str = "Sheet1",
        headers: list[str] | None = None,
        start_cell: str = "A1",
        overwrite: bool = False,
        auto_format: bool = True,
    ) -> dict[str, Any]:
        """
        Write data to an Excel file.

        Creates a new Excel file with the specified data. If headers are
        provided, they will be written first with bold formatting.

        Args:
            file_path: Path where the file will be written.
            rows: List of rows, where each row is a list of values.
            sheet_name: Name of the sheet. Defaults to "Sheet1".
            headers: Optional list of column headers.
            start_cell: Starting cell for data (A1 notation). Defaults to "A1".
            overwrite: Whether to overwrite an existing file.
            auto_format: Whether to auto-format column widths.

        Returns:
            Dictionary containing:
                - file_path: Path to the written file
                - rows_written: Number of data rows written
                - file_size_bytes: Size of the file in bytes

        Raises:
            WriteError: If writing fails.
            ExcelPermissionError: If the file cannot be written due to permissions.
        """
        path = self._validate_output_path(file_path, overwrite)

        try:
            workbook = Workbook()
            # Remove default sheet if it exists and has a different name
            if workbook.active is not None:
                default_sheet = workbook.active
                if default_sheet.title != sheet_name:
                    workbook.remove(default_sheet)

            worksheet = workbook.create_sheet(sheet_name) if sheet_name not in workbook.sheetnames else workbook[sheet_name]

            # Header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            start_row, start_col = self._parse_start_cell(start_cell)

            current_row = start_row + 1  # openpyxl uses 1-based indexing
            rows_written = 0

            if headers:
                for col_idx, header in enumerate(headers):
                    cell = worksheet.cell(
                        row=current_row,
                        column=start_col + col_idx + 1,  # 1-based
                        value=header,
                    )
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = header_border
                current_row += 1

            for row_data in rows:
                for col_idx, value in enumerate(row_data):
                    worksheet.cell(
                        row=current_row,
                        column=start_col + col_idx + 1,  # 1-based
                        value=value,
                    )
                current_row += 1
                rows_written += 1

            if auto_format:
                column_widths = self._calculate_column_widths(rows, headers)
                for col_idx, width in enumerate(column_widths):
                    col_letter = get_column_letter(start_col + col_idx + 1)
                    worksheet.column_dimensions[col_letter].width = width

            workbook.save(str(path))
            workbook.close()

            file_size = path.stat().st_size if path.exists() else 0

            return {
                "file_path": str(path.absolute()),
                "rows_written": rows_written,
                "file_size_bytes": file_size,
            }

        except PermissionError as e:
            raise ExcelPermissionError(
                file_path=file_path,
                operation="write",
            ) from e
        except Exception as e:
            raise WriteError(
                file_path=file_path,
                operation="write",
                reason=str(e),
            ) from e

    def write_multiple_sheets(
        self,
        file_path: str,
        sheets_data: dict[str, dict[str, Any]],
        overwrite: bool = False,
        auto_format: bool = True,
    ) -> dict[str, Any]:
        """
        Write multiple sheets to an Excel file.

        Creates a new Excel file with multiple sheets. Each sheet is specified
        by a dictionary entry with sheet name as key and configuration as value.

        Args:
            file_path: Path where the file will be written.
            sheets_data: Dictionary mapping sheet names to sheet configurations.
                Each configuration should contain:
                    - rows: List of data rows (required)
                    - headers: Optional list of headers
                    - start_cell: Starting cell (default: "A1")
            overwrite: Whether to overwrite an existing file.
            auto_format: Whether to auto-format column widths.

        Returns:
            Dictionary containing:
                - file_path: Path to the written file
                - sheets_written: Number of sheets written
                - total_rows_written: Total number of data rows written
                - file_size_bytes: Size of the file in bytes

        Raises:
            WriteError: If writing fails.
            ExcelPermissionError: If the file cannot be written due to permissions.
        """
        path = self._validate_output_path(file_path, overwrite)

        try:
            workbook = Workbook()
            # Remove the default sheet
            if workbook.active is not None:
                workbook.remove(workbook.active)

            # Header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            total_rows_written = 0
            sheets_written = 0

            for sheet_name, config in sheets_data.items():
                rows = config.get("rows", [])
                headers = config.get("headers")
                start_cell = config.get("start_cell", "A1")

                worksheet = workbook.create_sheet(sheet_name)

                start_row, start_col = self._parse_start_cell(start_cell)
                current_row = start_row + 1  # openpyxl uses 1-based indexing

                if headers:
                    for col_idx, header in enumerate(headers):
                        cell = worksheet.cell(
                            row=current_row,
                            column=start_col + col_idx + 1,
                            value=header,
                        )
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.border = header_border
                    current_row += 1

                for row_data in rows:
                    for col_idx, value in enumerate(row_data):
                        worksheet.cell(
                            row=current_row,
                            column=start_col + col_idx + 1,
                            value=value,
                        )
                    current_row += 1
                    total_rows_written += 1

                if auto_format:
                    column_widths = self._calculate_column_widths(rows, headers)
                    for col_idx, width in enumerate(column_widths):
                        col_letter = get_column_letter(start_col + col_idx + 1)
                        worksheet.column_dimensions[col_letter].width = width

                sheets_written += 1

            workbook.save(str(path))
            workbook.close()

            file_size = path.stat().st_size if path.exists() else 0

            return {
                "file_path": str(path.absolute()),
                "sheets_written": sheets_written,
                "total_rows_written": total_rows_written,
                "file_size_bytes": file_size,
            }

        except PermissionError as e:
            raise ExcelPermissionError(
                file_path=file_path,
                operation="write",
            ) from e
        except Exception as e:
            raise WriteError(
                file_path=file_path,
                operation="write",
                reason=str(e),
            ) from e

    def modify_existing_workbook(
        self,
        file_path: str,
        sheet_name: str,
        rows: list[list[Any]],
        start_cell: str = "A1",
        create_sheet_if_missing: bool = True,
    ) -> dict[str, Any]:
        """
        Modify an existing Excel workbook by adding/updating data.

        This method is unique to openpyxl adapter - it allows modifying
        existing workbooks while preserving other sheets and formatting.

        Args:
            file_path: Path to the existing Excel file.
            sheet_name: Name of the sheet to modify or create.
            rows: List of rows to write.
            start_cell: Starting cell for data (A1 notation). Defaults to "A1".
            create_sheet_if_missing: Whether to create the sheet if it doesn't exist.

        Returns:
            Dictionary containing:
                - file_path: Path to the modified file
                - rows_written: Number of data rows written
                - sheet_created: Whether a new sheet was created

        Raises:
            ExcelFileNotFoundError: If the file does not exist.
            InvalidFileFormatError: If the file format is not supported.
            SheetNotFoundError: If the sheet doesn't exist and create_sheet_if_missing is False.
            WriteError: If writing fails.
        """
        path = self._validate_file_path(file_path)

        try:
            workbook = load_workbook(str(path))

            sheet_created = False
            if sheet_name not in workbook.sheetnames:
                if create_sheet_if_missing:
                    workbook.create_sheet(sheet_name)
                    sheet_created = True
                else:
                    workbook.close()
                    raise SheetNotFoundError(
                        sheet_name=sheet_name,
                        available_sheets=workbook.sheetnames,
                    )

            worksheet = workbook[sheet_name]

            start_row, start_col = self._parse_start_cell(start_cell)
            current_row = start_row + 1  # openpyxl uses 1-based indexing
            rows_written = 0

            for row_data in rows:
                for col_idx, value in enumerate(row_data):
                    worksheet.cell(
                        row=current_row,
                        column=start_col + col_idx + 1,
                        value=value,
                    )
                current_row += 1
                rows_written += 1

            workbook.save(str(path))
            workbook.close()

            return {
                "file_path": str(path.absolute()),
                "rows_written": rows_written,
                "sheet_created": sheet_created,
            }

        except SheetNotFoundError:
            raise
        except PermissionError as e:
            raise ExcelPermissionError(
                file_path=file_path,
                operation="write",
            ) from e
        except Exception as e:
            raise WriteError(
                file_path=file_path,
                operation="modify",
                reason=str(e),
            ) from e
